import openpyxl

# ファイル名を指定する
txt_filename  = 'sample.txt'
xlsx_filename = 'sample.xlsx'

# 既存のEXCELファイルを開く
workbook = openpyxl.load_workbook(xlsx_filename)
sheet = workbook.active

# テキストファイルを行ごとに読み込み、EXCELワークシートに書き込む
with open(txt_filename, 'r') as file:

    # EXCELワークシートの2行目から書き込む
    row_num = 2

    for line in file:

        # 行をタブで分割して、3つのフィールドに分ける
        fields = line.strip().split('\t')

        # フィールドをA列、B列、C列に追加する
        sheet.cell(row=row_num, column=1, value=fields[0])
        sheet.cell(row=row_num, column=2, value=fields[1])
        sheet.cell(row=row_num, column=3, value=fields[2])

        row_num += 1

# EXCELファイルを保存し、閉じる
workbook.save(xlsx_filename)
workbook.close()